from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel
import io
import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models import SafetyIssue, ExportTemplate
from app.schemas import TemplateCreate, TemplateUpdate, TemplateResponse

router = APIRouter(prefix="/api/export", tags=["数据导出"])

# 对齐方式（供多处复用）
center_alignment = Alignment(horizontal="center", vertical="center")
left_alignment = Alignment(horizontal="left", vertical="center")
top_left_alignment = Alignment(horizontal="left", vertical="top")


def _get_or_create_default_templates(db: Session) -> List[dict]:
    """确保数据库里有默认模板，返回列表"""
    existing = db.query(ExportTemplate).all()
    if existing:
        return [t.to_dict() for t in existing]

    defaults = [
        {
            "name": "隐患检查表",
            "title_format": "",
            "columns": ["序号", "现场图片", "检查发现的主要隐患或问题", "法规名称、代码和条款号", "整改措施或建议", "备注"],
            "is_default": True,
        },
        {
            "name": "整改回复报告",
            "title_format": "《关于{date}安全隐患整改有关事项回复》",
            "columns": [],
            "is_default": False,
        },
    ]
    for d in defaults:
        t = ExportTemplate(
            name=d["name"],
            title_format=d["title_format"],
            columns_config=json.dumps(d["columns"]),
            is_default=d["is_default"],
        )
        db.add(t)
    db.commit()
    return db.query(ExportTemplate).all()


@router.get("/templates", response_model=List[TemplateResponse])
def get_templates(db: Session = Depends(get_db)):
    """获取所有导出模板"""
    templates = _get_or_create_default_templates(db)
    return [t.to_dict() if isinstance(t, ExportTemplate) else t for t in templates]


@router.post("/templates", response_model=TemplateResponse)
def create_template(template: TemplateCreate, db: Session = Depends(get_db)):
    """创建新模板"""
    if template.is_default:
        db.query(ExportTemplate).update({"is_default": False})

    new_template = ExportTemplate(
        name=template.name,
        title_format=template.title_format,
        columns_config=json.dumps(template.columns),
        is_default=template.is_default,
    )
    db.add(new_template)
    db.commit()
    db.refresh(new_template)
    return new_template.to_dict()


@router.put("/templates/{template_id}", response_model=TemplateResponse)
def update_template(template_id: int, template: TemplateUpdate, db: Session = Depends(get_db)):
    """更新模板"""
    db_template = db.query(ExportTemplate).filter(ExportTemplate.id == template_id).first()
    if not db_template:
        raise HTTPException(status_code=404, detail="模板不存在")

    if template.is_default:
        db.query(ExportTemplate).update({"is_default": False})

    update_data = template.dict(exclude_unset=True)
    if "columns" in update_data:
        update_data["columns_config"] = json.dumps(update_data.pop("columns"))
    for key, value in update_data.items():
        setattr(db_template, key, value)

    db.commit()
    db.refresh(db_template)
    return db_template.to_dict()


@router.delete("/templates/{template_id}")
def delete_template(template_id: int, db: Session = Depends(get_db)):
    """删除模板（禁止删除默认模板）"""
    db_template = db.query(ExportTemplate).filter(ExportTemplate.id == template_id).first()
    if not db_template:
        raise HTTPException(status_code=404, detail="模板不存在")
    if db_template.is_default:
        raise HTTPException(status_code=400, detail="默认模板不能删除")
    db.delete(db_template)
    db.commit()
    return {"message": "删除成功"}


# ─────────────────────────────────────────────────────────────────────────────
#  导出 Excel（隐患检查表）
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/excel")
def export_to_excel(
    status: str = None,
    severity: str = None,
    template_id: int = 1,
    db: Session = Depends(get_db)
):
    """
    按隐患检查表模板导出 Excel。
    模板字段与 /api/export/templates 中 id=1（隐患检查表）对齐。
    """
    query = db.query(SafetyIssue)
    if status:
        query = query.filter(SafetyIssue.status == status)
    if severity:
        query = query.filter(SafetyIssue.severity == severity)

    issues = query.order_by(SafetyIssue.create_time.desc()).all()
    if not issues:
        raise HTTPException(status_code=404, detail="没有可导出的问题")

    wb = Workbook()
    ws = wb.active
    ws.title = "隐患检查表"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    normal_font = Font(size=11)

    column_widths = {'A': 8, 'B': 30, 'C': 40, 'D': 50, 'E': 40, 'F': 20}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    project_name = issues[0].project_name if issues[0].project_name else "未填写"
    check_time = datetime.now().strftime("%Y.%m.%d")

    # 第1行：标题
    ws.merge_cells('A1:F1')
    c = ws.cell(row=1, column=1, value="委托项目安全生产隐患检查表")
    c.font = title_font
    c.alignment = center_alignment
    ws.row_dimensions[1].height = 35

    # 第2行：企业信息
    ws.merge_cells('A2:F2')
    info = f"企业名称：    {project_name}              隐患条数：  {len(issues)}  条                      检查时间：  {check_time}"
    c = ws.cell(row=2, column=1, value=info)
    c.font = normal_font
    ws.row_dimensions[2].height = 25

    # 第3行：说明
    ws.merge_cells('A3:F3')
    note = ("注：1.发现的主要内容或问题应配有图片和文字描述（时间、地点（位置）、现状或情形）；"
            "2.表中\"法规\"泛指国家法律法规、规章、标准和规范以及被检查企业安全生产管理规章制度、操作规程等；"
            "3.表中\"整改措施或建议\"可包括整改措施、时限和预案等内容。")
    c = ws.cell(row=3, column=1, value=note)
    c.font = Font(size=10)
    ws.row_dimensions[3].height = 40

    # 第4行起：表头
    headers = ["序号", "现场图片", "检查发现的主要隐患或问题", "法规名称、代码和条款号", "整改措施或建议", "备注"]
    for col_num, header in enumerate(headers, 1):
        c = ws.cell(row=4, column=col_num, value=header)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center_alignment
        c.border = thin_border
    ws.row_dimensions[4].height = 25

    current_row = 5
    for idx, issue in enumerate(issues, 1):
        ws.cell(row=current_row, column=1, value=idx).border = thin_border
        ws.cell(row=current_row, column=1).alignment = center_alignment

        issue_photos = [p for p in issue.photos if p.photo_type == "问题照片"]
        if issue_photos:
            photo = issue_photos[0]
            try:
                if os.path.exists(photo.file_path):
                    img = XLImage(photo.file_path)
                    img.width = 120
                    img.height = 90
                    ws.add_image(img, f'B{current_row}')
            except Exception:
                pass

        ws.cell(row=current_row, column=2, value="").border = thin_border
        ws.cell(row=current_row, column=3, value=issue.title).border = thin_border
        ws.cell(row=current_row, column=4, value=issue.description or "").border = thin_border
        ws.cell(row=current_row, column=5, value=issue.notes or "").border = thin_border

        deadline_str = ""
        if issue.deadline:
            deadline_str = f"整改时限：{issue.deadline.month}月{issue.deadline.day}日"
        ws.cell(row=current_row, column=6, value=deadline_str).border = thin_border
        ws.row_dimensions[current_row].height = 100
        current_row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"safety_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ─────────────────────────────────────────────────────────────────────────────
#  导出 Excel（带照片，完整版）
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/excel-with-photos")
def export_excel_with_photos(
    status: str = None,
    severity: str = None,
    db: Session = Depends(get_db)
):
    """导出带所有照片的安全问题 Excel"""
    query = db.query(SafetyIssue)
    if status:
        query = query.filter(SafetyIssue.status == status)
    if severity:
        query = query.filter(SafetyIssue.severity == severity)

    issues = query.order_by(SafetyIssue.create_time.desc()).all()
    if not issues:
        raise HTTPException(status_code=404, detail="没有可导出的问题")

    wb = Workbook()
    ws = wb.active
    ws.title = "安全问题"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for col, width in [('A', 8), ('B', 30), ('C', 40), ('D', 50), ('E', 40), ('F', 20)]:
        ws.column_dimensions[col].width = width

    headers = ["序号", "现场图片", "检查发现的主要隐患或问题", "法规名称、代码和条款号", "整改措施或建议", "备注"]
    for col_num, header in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_num, value=header)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center_alignment
        c.border = thin_border

    current_row = 2
    for idx, issue in enumerate(issues, 1):
        ws.cell(row=current_row, column=1, value=idx).border = thin_border
        ws.row_dimensions[current_row].height = 120

        issue_photos = [p for p in issue.photos if p.photo_type == "问题照片"]
        for photo_idx, photo in enumerate(issue_photos):
            try:
                if os.path.exists(photo.file_path):
                    img = XLImage(photo.file_path)
                    img.width = 150
                    img.height = 120
                    ws.add_image(img, f'B{current_row}')
                    break
            except Exception:
                pass

        ws.cell(row=current_row, column=3, value=issue.title).border = thin_border
        ws.cell(row=current_row, column=4, value=issue.description or "").border = thin_border
        ws.cell(row=current_row, column=5, value=issue.notes or "").border = thin_border

        remark_parts = []
        if issue.severity:
            remark_parts.append(f"严重程度: {issue.severity}")
        if issue.deadline:
            remark_parts.append(f"整改时限: {issue.deadline.isoformat()}")
        if issue.status:
            remark_parts.append(f"状态: {issue.status}")
        if issue.responsible_person:
            remark_parts.append(f"责任人: {issue.responsible_person}")
        if issue.location:
            remark_parts.append(f"位置: {issue.location}")
        ws.cell(row=current_row, column=6, value="; ".join(remark_parts)).border = thin_border
        current_row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"safety_report_with_photos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ─────────────────────────────────────────────────────────────────────────────
#  整改回复报告（POST，支持选择隐患）
# ─────────────────────────────────────────────────────────────────────────────

class RectificationReplyRequest(BaseModel):
    project_name: Optional[str] = None  # 不填则自动取隐患列表第一个
    project_responsible: str
    reply_date: str
    issue_ids: Optional[List[int]] = None


@router.post("/rectification-reply")
def export_rectification_reply(
    request: RectificationReplyRequest,
    db: Session = Depends(get_db)
):
    """
    按标准整改回复格式导出 Excel。
    格式：每条隐患占 4 行，结构如下：
      行(4n)   : 标题 + 隐患事项n（合并 B:E）
      行(4n+1) : 整改措施（合并 B:E）
      行(4n+2) : 整改前/后照片标签
      行(4n+3) : 照片区域（合并 B:C / D:E）
    末尾：回复日期行
    列宽：A=14.875, B=9.625, C=25.625, D=9.625, E=25.625
    """
    query = db.query(SafetyIssue)
    if request.issue_ids:
        query = query.filter(SafetyIssue.id.in_(request.issue_ids))
    issues = query.order_by(SafetyIssue.create_time.desc()).all()
    if not issues:
        raise HTTPException(status_code=404, detail="没有可导出的问题")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 列宽（与标准模板一致）
    ws.column_dimensions['A'].width = 14.875
    ws.column_dimensions['B'].width = 9.625
    ws.column_dimensions['C'].width = 25.625
    ws.column_dimensions['D'].width = 9.625
    ws.column_dimensions['E'].width = 25.625

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=11)
    normal_font = Font(size=11)

    def set_border_row(row, cols=range(1, 6)):
        for col in cols:
            ws.cell(row=row, column=col).border = thin_border

    current_row = 1

    # ── 第1行：标题 ──────────────────────────────────────────────────────────
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws.cell(row=current_row, column=1,
            value=f"《关于{request.reply_date}安全隐患整改有关事项回复》")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).alignment = center_alignment
    ws.row_dimensions[current_row].height = 30
    set_border_row(current_row)
    current_row += 1

    # ── 第2行：项目名称 ──────────────────────────────────────────────────────
    ws.cell(row=current_row, column=1, value="项目名称").font = header_font
    ws.cell(row=current_row, column=1).alignment = left_alignment
    set_border_row(current_row)
    ws.merge_cells(f'B{current_row}:E{current_row}')
    ws.cell(row=current_row, column=2, value=request.project_name).font = normal_font
    ws.cell(row=current_row, column=2).alignment = left_alignment
    for col in range(2, 6):
        ws.cell(row=current_row, column=col).border = thin_border
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # ── 第3行：项目负责人 ────────────────────────────────────────────────────
    ws.cell(row=current_row, column=1, value="项目负责人").font = header_font
    ws.cell(row=current_row, column=1).alignment = left_alignment
    set_border_row(current_row)
    ws.merge_cells(f'B{current_row}:E{current_row}')
    ws.cell(row=current_row, column=2, value=request.project_responsible).font = normal_font
    ws.cell(row=current_row, column=2).alignment = left_alignment
    for col in range(2, 6):
        ws.cell(row=current_row, column=col).border = thin_border
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # ── 每条隐患：4行 ────────────────────────────────────────────────────────
    for idx, issue in enumerate(issues, 1):
        issue_top = current_row

        # 行 A：整改事项回复（合并 A 列跨 4 行）
        ws.merge_cells(f'A{issue_top}:A{issue_top + 3}')
        c = ws.cell(row=issue_top, column=1, value="整改事项回复")
        c.font = header_font
        c.alignment = center_alignment
        c.border = thin_border
        # 其余 A 列行也要加边框
        for r in range(issue_top + 1, issue_top + 4):
            ws.cell(row=r, column=1).border = thin_border

        # 行 B：隐患标题（合并 B:E）
        ws.merge_cells(f'B{issue_top}:E{issue_top}')
        c = ws.cell(row=issue_top, column=2, value=f"隐患事项{idx}：{issue.title}")
        c.font = header_font
        c.alignment = left_alignment
        for col in range(2, 6):
            ws.cell(row=issue_top, column=col).border = thin_border
        ws.row_dimensions[issue_top].height = 32
        current_row = issue_top + 1

        # 行 C：整改措施（合并 B:E）
        ws.merge_cells(f'B{current_row}:E{current_row}')
        c = ws.cell(row=current_row, column=2,
                    value=f"整改措施：{issue.notes if issue.notes else '已整改'}")
        c.font = normal_font
        c.alignment = left_alignment
        for col in range(2, 6):
            ws.cell(row=current_row, column=col).border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # 行 D：照片（合并 B:C / D:E），标签写在格子的左上角
        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws.merge_cells(f'D{current_row}:E{current_row}')
        for col in range(1, 6):
            ws.cell(row=current_row, column=col).border = thin_border
        # 照片格左上角写标签
        cell_b = ws.cell(row=current_row, column=2, value="整改前照片：")
        cell_b.font = header_font
        cell_b.alignment = top_left_alignment
        cell_d = ws.cell(row=current_row, column=4, value="整改后照片：")
        cell_d.font = header_font
        cell_d.alignment = top_left_alignment

        issue_photos = [p for p in issue.photos if p.photo_type == "问题照片"]
        rect_photos = [p for p in issue.photos if p.photo_type == "整改照片"]

        if issue_photos:
            try:
                if os.path.exists(issue_photos[0].file_path):
                    img = XLImage(issue_photos[0].file_path)
                    img.width = 150
                    img.height = 100
                    ws.add_image(img, f'B{current_row}')
            except Exception:
                pass

        if rect_photos:
            try:
                if os.path.exists(rect_photos[0].file_path):
                    img = XLImage(rect_photos[0].file_path)
                    img.width = 150
                    img.height = 100
                    ws.add_image(img, f'D{current_row}')
            except Exception:
                pass

        ws.row_dimensions[current_row].height = 106
        current_row += 1

    # ── 最后：回复日期 ────────────────────────────────────────────────────────
    ws.cell(row=current_row, column=1, value="回复日期").font = header_font
    ws.cell(row=current_row, column=1).alignment = left_alignment
    set_border_row(current_row)
    ws.merge_cells(f'B{current_row}:E{current_row}')
    ws.cell(row=current_row, column=2, value=request.reply_date).font = normal_font
    ws.cell(row=current_row, column=2).alignment = left_alignment
    for col in range(2, 6):
        ws.cell(row=current_row, column=col).border = thin_border
    ws.row_dimensions[current_row].height = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"rectification_reply_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ─────────────────────────────────────────────────────────────────────────────
#  旧版 GET 导出（保留兼容）
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/excel/rectification")
def export_rectification_report(
    status: str = None,
    project_name: str = None,
    responsible_person: str = None,
    db: Session = Depends(get_db)
):
    """旧版 GET 整改回复（兼容保留）"""
    query = db.query(SafetyIssue)
    if status:
        query = query.filter(SafetyIssue.status == status)
    if project_name:
        query = query.filter(SafetyIssue.project_name.like(f"%{project_name}%"))
    issues = query.order_by(SafetyIssue.create_time.desc()).all()
    if not issues:
        raise HTTPException(status_code=404, detail="没有可导出的问题")

    wb = Workbook()
    ws = wb.active
    ws.title = "整改回复"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80

    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=11)

    current_row = 1
    today = datetime.now().strftime("%Y年%m月%d日")

    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws.cell(row=current_row, column=1, value=f"《关于{today}安全隐患整改有关事项回复》")
    ws.cell(row=current_row, column=1).font = title_font
    ws.cell(row=current_row, column=1).alignment = center_alignment
    ws.row_dimensions[current_row].height = 35
    current_row += 1

    def add_info_row(label, value):
        nonlocal current_row
        ws.cell(row=current_row, column=1, value=label).font = header_font
        ws.cell(row=current_row, column=2, value=value)
        ws.row_dimensions[current_row].height = 25
        for col in range(1, 3):
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1

    add_info_row("项目名称", project_name or issues[0].project_name or "请填写项目名称")
    add_info_row("项目负责人", responsible_person or issues[0].responsible_person or "请填写负责人")

    for idx, issue in enumerate(issues, 1):
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws.cell(row=current_row, column=1, value=f"隐患事项{idx}：{issue.title}").font = header_font
        ws.row_dimensions[current_row].height = 25
        for col in range(1, 3):
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1

        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws.cell(row=current_row, column=1, value=f"整改措施：{issue.notes if issue.notes else '已整改'}")
        ws.row_dimensions[current_row].height = 25
        for col in range(1, 3):
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1

        issue_photos = [p for p in issue.photos if p.photo_type == "问题照片"]
        rect_photos = [p for p in issue.photos if p.photo_type == "整改照片"]
        photo_height = 150

        for photo_idx in range(max(len(issue_photos), len(rect_photos), 1)):
            c1 = ws.cell(row=current_row, column=1, value="整改前照片：")
            c1.font = header_font
            c1.alignment = top_left_alignment
            c2 = ws.cell(row=current_row, column=2, value="整改后照片：")
            c2.font = header_font
            c2.alignment = top_left_alignment
            if photo_idx < len(issue_photos):
                try:
                    if os.path.exists(issue_photos[photo_idx].file_path):
                        img = XLImage(issue_photos[photo_idx].file_path)
                        img.width = 200
                        img.height = 150
                        ws.add_image(img, f'A{current_row}')
                except Exception:
                    pass
            if photo_idx < len(rect_photos):
                try:
                    if os.path.exists(rect_photos[photo_idx].file_path):
                        img = XLImage(rect_photos[photo_idx].file_path)
                        img.width = 200
                        img.height = 150
                        ws.add_image(img, f'B{current_row}')
                except Exception:
                    pass
            ws.row_dimensions[current_row].height = photo_height
            for col in range(1, 3):
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws.cell(row=current_row, column=1, value="整改事项回复")
        ws.row_dimensions[current_row].height = photo_height
        for col in range(1, 3):
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1
        current_row += 1

    add_info_row("回复日期", datetime.now().strftime("%Y年%m月%d日"))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"rectification_reply_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
