import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('仙境花园整改回复.xlsx')
ws = wb.active

print('=== 工作表信息 ===')
print(f'工作表名称: {ws.title}')
print(f'最大行: {ws.max_row}')
print(f'最大列: {ws.max_column}')

print('\n=== 列宽信息 ===')
for col in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col)
    if col_letter in ws.column_dimensions:
        print(f'{col_letter}: {ws.column_dimensions[col_letter].width}')

print('\n=== 内容预览 ===')
for row in range(1, min(ws.max_row + 1, 50)):
    row_data = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        value = str(cell.value) if cell.value else ''
        row_data.append(value[:80] if len(value) > 80 else value)
    print(f'行{row}: {row_data}')

print('\n=== 合并单元格 ===')
for merged_range in ws.merged_cells.ranges:
    print(f'{merged_range}')

print('\n=== 行高信息 ===')
for row in range(1, min(ws.max_row + 1, 20)):
    if row in ws.row_dimensions:
        print(f'行{row}: {ws.row_dimensions[row].height}')

print('\n=== 边框信息 ===')
thin_border_count = 0
for row in range(1, min(ws.max_row + 1, 20)):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.border and cell.border.left and cell.border.left.style:
            thin_border_count += 1
print(f'有边框的单元格数量: {thin_border_count}')